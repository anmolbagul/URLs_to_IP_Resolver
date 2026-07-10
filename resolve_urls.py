#!/usr/bin/env python3

import socket
import argparse
from urllib.parse import urlparse
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import PatternFill


DUPLICATE_COLORS = [
    "FFFF00",  # Yellow
    "90EE90",  # Light Green
    "ADD8E6",  # Light Blue
    "FFA07A",  # Light Salmon
    "DDA0DD",  # Plum
    "F4A460",  # Sandy Brown
    "AFEEEE",  # Pale Turquoise
    "FFC0CB",  # Pink
]


def extract_hostname(url):
    url = url.strip()

    if not url:
        return None

    if not url.startswith(("http://", "https://")):
        url = "http://" + url

    parsed = urlparse(url)
    hostname = parsed.hostname

    return hostname


def resolve_hostname(hostname):
    try:
        ip_list = socket.gethostbyname_ex(hostname)[2]
        return sorted(set(ip_list))
    except Exception:
        return []


def main(input_file, output_file):
    results = []
    ip_to_rows = defaultdict(list)

    with open(input_file, "r", encoding="utf-8") as file:
        urls = file.readlines()

    for url in urls:
        original_url = url.strip()
        hostname = extract_hostname(original_url)

        if not hostname:
            continue

        ips = resolve_hostname(hostname)

        if ips:
            for ip in ips:
                results.append([original_url, hostname, ip])
        else:
            results.append([original_url, hostname, "Could not resolve"])

    wb = Workbook()
    ws = wb.active
    ws.title = "NSLookup Results"

    headers = ["Original URL", "Hostname", "IP Address"]
    ws.append(headers)

    for row in results:
        ws.append(row)

    for row_num in range(2, ws.max_row + 1):
        ip = ws.cell(row=row_num, column=3).value
        if ip and ip != "Could not resolve":
            ip_to_rows[ip].append(row_num)

    color_index = 0

    for ip, rows in ip_to_rows.items():
        if len(rows) > 1:
            fill = PatternFill(
                start_color=DUPLICATE_COLORS[color_index % len(DUPLICATE_COLORS)],
                end_color=DUPLICATE_COLORS[color_index % len(DUPLICATE_COLORS)],
                fill_type="solid"
            )

            for row_num in rows:
                for col_num in range(1, 4):
                    ws.cell(row=row_num, column=col_num).fill = fill

            color_index += 1

    for column_cells in ws.columns:
        max_length = max(len(str(cell.value)) if cell.value else 0 for cell in column_cells)
        ws.column_dimensions[column_cells[0].column_letter].width = max_length + 3

    wb.save(output_file)
    print(f"[+] Results saved to: {output_file}")


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Resolve hostnames from URLs and highlight duplicate IPs.")
    parser.add_argument("-i", "--input", required=True, help="Input TXT file containing URLs")
    parser.add_argument("-o", "--output", default="nslookup_results.xlsx", help="Output Excel file")

    args = parser.parse_args()
    main(args.input, args.output)